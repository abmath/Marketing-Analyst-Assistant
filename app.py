"""
Marketing Analytics Assistant
Adobe Analytics Excel dumps → NL queries → PowerPoint insights
Backend: Ollama + Microsoft Phi-3
"""

import re
import traceback
from io import BytesIO

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook

# ── Config ─────────────────────────────────────────────────────────────────────
OLLAMA_BASE = "http://localhost:11434"
DEFAULT_MODEL = "phi3"

# ── Page setup ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Marketing Analytics Assistant",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  .block-container { padding-top: 1.5rem; }
  h1 { font-size: 1.9rem !important; }
  .section-label {
    font-size: 1.05rem; font-weight: 700;
    color: #0f3460; border-bottom: 2px solid #0f3460;
    padding-bottom: 4px; margin-bottom: 12px;
  }
  .insight-card {
    background: #f5f7ff; border-left: 4px solid #0f3460;
    padding: 14px 16px; margin: 10px 0; border-radius: 6px;
    white-space: pre-wrap; font-size: 0.92rem;
  }
  .stAlert { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)


# ── Ollama helpers ─────────────────────────────────────────────────────────────
@st.cache_data(ttl=10)
def check_ollama():
    try:
        r = requests.get(f"{OLLAMA_BASE}/api/tags", timeout=3)
        models = [m["name"] for m in r.json().get("models", [])]
        return True, models
    except Exception:
        return False, []


def ollama_generate(prompt: str, model: str, temperature: float = 0.1) -> str:
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": temperature},
    }
    r = requests.post(f"{OLLAMA_BASE}/api/generate", json=payload, timeout=180)
    r.raise_for_status()
    return r.json().get("response", "")


# ── Excel parsing ──────────────────────────────────────────────────────────────
def split_into_segments(ws):
    """
    Read all rows from a worksheet and split on fully-blank rows.
    Returns list of row-lists (each list is a contiguous block).
    """
    segments, current = [], []
    for row in ws.iter_rows(values_only=True):
        non_empty = [v for v in row if v is not None and str(v).strip() != ""]
        if non_empty:
            current.append(list(row))
        else:
            if len(current) >= 2:
                segments.append(current)
            current = []
    if len(current) >= 2:
        segments.append(current)
    return segments


def rows_to_df(rows: list) -> pd.DataFrame:
    headers = [str(h) if h is not None else f"Col_{i}" for i, h in enumerate(rows[0])]
    df = pd.DataFrame(rows[1:], columns=headers)
    df = df.dropna(how="all").reset_index(drop=True)
    # Drop fully-null columns
    df = df.loc[:, df.notna().any()]
    # Infer better dtypes
    df = df.infer_objects()
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col])
        except (ValueError, TypeError):
            pass
    return df


@st.cache_data(show_spinner=False)
def parse_excel(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    tables: dict[str, pd.DataFrame] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 1 — Named Excel Table objects (highest fidelity)
        sheet_tables = list(ws.tables.values())
        if sheet_tables:
            for tbl in sheet_tables:
                try:
                    rows = [[cell.value for cell in row] for row in ws[tbl.ref]]
                    if len(rows) >= 2:
                        df = rows_to_df(rows)
                        label = f"{sheet_name} › {tbl.displayName or tbl.name}"
                        tables[label] = df
                except Exception:
                    pass
            continue

        # 2 — Auto-detect sub-tables separated by blank rows
        segments = split_into_segments(ws)

        if len(segments) > 1:
            for seg in segments:
                # Use first cell of first row as a human-readable label
                raw_label = seg[0][0]
                label_str = str(raw_label).strip() if raw_label else ""
                label = f"{sheet_name} › {label_str}" if label_str else f"{sheet_name} › table"
                # Make unique
                base = label
                counter = 2
                while label in tables:
                    label = f"{base} ({counter})"
                    counter += 1
                df = rows_to_df(seg)
                if not df.empty:
                    tables[label] = df
        elif segments:
            df = rows_to_df(segments[0])
            if not df.empty:
                tables[sheet_name] = df

    return tables


# ── NL query engine ────────────────────────────────────────────────────────────
def _df_schema(df: pd.DataFrame) -> str:
    """Build a schema block with exact column names, types, and sample values."""
    num_cols  = df.select_dtypes("number").columns.tolist()
    cat_cols  = df.select_dtypes("object").columns.tolist()
    date_cols = [c for c in df.columns if "date" in c.lower() or "week" in c.lower() or "month" in c.lower()]

    lines = [f"Shape: {len(df)} rows × {len(df.columns)} columns", ""]
    lines.append("# Numeric columns (use these for calculations / outlier detection):")
    for c in num_cols[:20]:
        s = df[c].dropna()
        lines.append(
            f"  df[{c!r}]  "
            f"min={s.min():.2f}  max={s.max():.2f}  "
            f"mean={s.mean():.2f}  std={s.std():.2f}"
        )
    lines.append("")
    lines.append("# Categorical columns:")
    for c in cat_cols[:10]:
        top = df[c].value_counts().head(3).index.tolist()
        lines.append(f"  df[{c!r}]  e.g. {top}")
    if date_cols:
        lines.append("")
        lines.append("# Date / time columns:")
        for c in date_cols[:5]:
            lines.append(f"  df[{c!r}]  e.g. {df[c].dropna().head(2).tolist()}")
    return "\n".join(lines)


def build_query_prompt(df: pd.DataFrame, question: str, error_feedback: str = "") -> str:
    schema = _df_schema(df)
    num_cols = df.select_dtypes("number").columns.tolist()

    error_block = ""
    if error_feedback:
        error_block = f"""
The previous attempt raised this error — fix it:
{error_feedback}
"""

    return f"""You are a Python/pandas data analyst. Write executable code to answer the question below.

=== DATAFRAME SCHEMA ===
{schema}

=== AVAILABLE VARIABLES (already in scope — do NOT import) ===
- df        : the pandas DataFrame above
- pd        : pandas
- np        : numpy
- num_cols  : {num_cols[:20]}   (list of numeric column names)

=== COMMON PATTERNS (use these as templates) ===
# Outlier detection via IQR:
Q1 = df['col'].quantile(0.25)
Q3 = df['col'].quantile(0.75)
IQR = Q3 - Q1
result = df[(df['col'] < Q1 - 1.5*IQR) | (df['col'] > Q3 + 1.5*IQR)]

# Z-score outliers across multiple numeric columns:
from_cols = num_cols[:5]
z = (df[from_cols] - df[from_cols].mean()) / df[from_cols].std()
result = df[(np.abs(z) > 2.5).any(axis=1)]

# Top/bottom N:
result = df.nlargest(10, 'col')

# Group aggregation:
result = df.groupby('Category')['Metric'].sum().reset_index().sort_values('Metric', ascending=False)
{error_block}
=== QUESTION ===
{question}

=== RULES ===
- Store the final answer in a variable called `result` (DataFrame, Series, scalar, or string)
- Use exact column names from the schema — copy them character-for-character including spaces and special chars
- Do NOT import anything; do NOT call print()
- Return ONLY a fenced python code block, nothing else

```python
"""


def extract_code(raw: str) -> str:
    # Try fenced block first
    m = re.search(r"```python\s*(.*?)```", raw, re.DOTALL)
    if m:
        return m.group(1).strip()
    # Model forgot closing fence — take everything after opening fence
    m2 = re.search(r"```python\s*(.*)", raw, re.DOTALL)
    if m2:
        return m2.group(1).strip()
    # No fence at all — return raw (might still work)
    return raw.strip()


def run_nl_query(df: pd.DataFrame, question: str, model: str, max_retries: int = 2):
    """
    Generate + execute pandas code. On failure, send the error back to the model
    for self-correction (up to max_retries attempts).
    """
    import numpy as np  # ensure available

    namespace = {"df": df.copy(), "pd": pd, "np": np,
                 "num_cols": df.select_dtypes("number").columns.tolist()}

    error_feedback = ""
    last_code = ""
    attempts = []

    for attempt in range(1, max_retries + 2):  # attempts: 1, 2, 3
        prompt = build_query_prompt(df, question, error_feedback)
        raw = ollama_generate(prompt, model=model, temperature=0.05 if attempt == 1 else 0.15)
        code = extract_code(raw)
        last_code = code
        attempts.append({"attempt": attempt, "code": code, "error": None})

        try:
            exec(compile(code, "<query>", "exec"), namespace)  # noqa: S102
            result = namespace.get("result")
            return result, attempts, None
        except Exception as exc:
            err_msg = f"{type(exc).__name__}: {exc}"
            attempts[-1]["error"] = err_msg
            error_feedback = f"Code that failed:\n```python\n{code}\n```\nError: {err_msg}"

    return None, attempts, attempts[-1]["error"]


# ── Insights engine ────────────────────────────────────────────────────────────
def build_insights_prompt(df: pd.DataFrame, table_name: str) -> str:
    numeric_summary = (
        df.describe().round(2).to_string()
        if not df.select_dtypes("number").empty
        else "(no numeric columns)"
    )
    cat_cols = df.select_dtypes("object").columns.tolist()[:5]
    cat_preview = ""
    for c in cat_cols:
        top = df[c].value_counts().head(5)
        cat_preview += f"\n  {c}: {top.to_dict()}"

    return f"""You are a marketing analytics consultant preparing a PowerPoint deck for the CMO.

Dataset: "{table_name}"
Rows: {len(df):,} | Columns: {list(df.columns[:20])}

Numeric summary:
{numeric_summary}

Top categorical values:{cat_preview if cat_preview else " (none)"}

Generate EXACTLY 5 PowerPoint slide suggestions. For each slide use this format:

## Slide N: [Title]
**Key message:** one sentence
**Bullets:**
- bullet 1
- bullet 2
- bullet 3
**Recommended chart:** bar / line / pie / scatter / heatmap

Focus on: performance trends, audience segments, conversion opportunities, anomalies, and strategic recommendations.
Be specific — reference column names and directional trends where possible.
"""


def get_insights(df: pd.DataFrame, table_name: str, model: str) -> str:
    return ollama_generate(
        build_insights_prompt(df, table_name), model=model, temperature=0.45
    )


# ── Sidebar ────────────────────────────────────────────────────────────────────
def render_sidebar():
    with st.sidebar:
        st.markdown("## ⚙️ Settings")
        ollama_ok, available_models = check_ollama()

        if ollama_ok:
            st.success("Ollama connected")
            phi_models = [m for m in available_models if "phi" in m.lower()]
            all_models = phi_models + [m for m in available_models if m not in phi_models]
            if all_models:
                model = st.selectbox("Model", all_models, index=0)
            else:
                st.warning("No models found.\nRun: `ollama pull phi3`")
                model = st.text_input("Model name", value=DEFAULT_MODEL)
        else:
            st.error("Ollama not reachable.\n\nStart with:\n```\nollama serve\n```")
            model = DEFAULT_MODEL

        st.divider()
        st.markdown("### 📂 Data")
        uploaded = st.file_uploader(
            "Upload Adobe Analytics export",
            type=["xlsx", "xls"],
            help="Supports multiple sheets and multiple tables per sheet",
        )

        st.divider()
        st.markdown(
            "**Phi-3** via [Ollama](https://ollama.com) · "
            "Pull with `ollama pull phi3`",
            unsafe_allow_html=False,
        )

    return uploaded, model, ollama_ok


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    uploaded, model, ollama_ok = render_sidebar()

    st.title("📊 Marketing Analytics Assistant")
    st.caption("Adobe Analytics · CRM · Website Traffic — powered by Phi-3 (Ollama)")

    if not uploaded:
        st.divider()
        c1, c2, c3 = st.columns(3)
        c1.info("**1 — Load**\nUpload any Adobe Analytics Excel export — multi-sheet, multi-table supported.")
        c2.info("**2 — Query**\nAsk questions in plain English. Phi-3 writes and runs the pandas code for you.")
        c3.info("**3 — Present**\nGet five PowerPoint slide suggestions with chart recommendations.")
        return

    # ── Parse Excel ────────────────────────────────────────────────────────────
    with st.spinner("Parsing Excel…"):
        file_bytes = uploaded.read()
        tables = parse_excel(file_bytes)

    if not tables:
        st.error("No tables found. Make sure the file has data rows with headers.")
        return

    st.divider()

    # ── Step 1: Table selection ────────────────────────────────────────────────
    st.markdown('<div class="section-label">Step 1 — Select a Table</div>', unsafe_allow_html=True)

    table_names = list(tables.keys())
    selected = st.selectbox(
        f"Found **{len(table_names)}** table(s) in this file:",
        table_names,
        format_func=lambda x: x,
    )
    df = tables[selected]

    with st.expander("🔍 Data Preview", expanded=True):
        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.metric("Rows", f"{len(df):,}")
        mc2.metric("Columns", len(df.columns))
        mc3.metric("Numeric cols", len(df.select_dtypes("number").columns))
        mc4.metric("Text cols", len(df.select_dtypes("object").columns))
        st.dataframe(df.head(100), use_container_width=True)

    st.divider()

    # ── Step 2: NL Query ───────────────────────────────────────────────────────
    st.markdown('<div class="section-label">Step 2 — Ask a Question</div>', unsafe_allow_html=True)

    example_questions = [
        "What are the top 10 pages by page views?",
        "Show conversion rate by traffic source",
        "Which campaign drove the most revenue?",
        "What is the average session duration by channel?",
        "Show me weekly trend of visits",
    ]

    left, right = st.columns([1, 2])
    with left:
        st.markdown("**Quick examples:**")
        for eq in example_questions:
            if st.button(eq, key=f"eq_{eq}", use_container_width=True):
                st.session_state["nl_q"] = eq

    with right:
        question = st.text_area(
            "Your question:",
            value=st.session_state.get("nl_q", ""),
            height=100,
            placeholder="e.g. Which pages have a bounce rate above 70%?",
            key="nl_q",
        )
        col_run, col_clear = st.columns([2, 1])
        run_clicked = col_run.button("▶ Run Query", type="primary", disabled=not ollama_ok)
        if col_clear.button("Clear"):
            st.session_state.pop("nl_q", None)
            st.rerun()

    if run_clicked and question.strip():
        with st.spinner(f"Asking {model}…"):
            result, attempts, error = run_nl_query(df, question.strip(), model)

        # Show code (and retry history if there were multiple attempts)
        with st.expander(
            f"🛠 Generated code {'(solved in 1 attempt)' if len(attempts) == 1 and not error else f'({len(attempts)} attempt(s))'}",
            expanded=False,
        ):
            for i, att in enumerate(attempts, 1):
                if len(attempts) > 1:
                    st.markdown(f"**Attempt {i}**")
                st.code(att["code"], language="python")
                if att["error"]:
                    st.caption(f"Error: {att['error']}")

        if error:
            st.error(f"**Could not execute after {len(attempts)} attempt(s):** {error}")
            st.caption(
                "Tip: rephrase using column names visible in the preview above, "
                "e.g. *'Show outliers in Bounce Rate (%)'* instead of *'bounce rate outliers'*."
            )
        elif result is None:
            st.warning("Query ran but returned `None`. Try adding 'show me' or 'return as a table'.")
        else:
            st.markdown("**Result:**")
            if isinstance(result, pd.DataFrame):
                st.dataframe(result, use_container_width=True)
                st.caption(f"{len(result):,} rows · {len(result.columns)} columns")

                # Auto chart if reasonable size
                res_num_cols = result.select_dtypes("number").columns.tolist()
                if 1 <= len(result) <= 60 and res_num_cols:
                    with st.expander("📈 Quick chart"):
                        chart_col = st.selectbox("Metric:", res_num_cols, key="chart_col")
                        index_col = result.columns[0] if result.columns[0] not in res_num_cols else None
                        chart_df = (
                            result.set_index(index_col)[chart_col]
                            if index_col else result[chart_col]
                        )
                        st.bar_chart(chart_df)

            elif isinstance(result, pd.Series):
                st.dataframe(result.rename("value").to_frame(), use_container_width=True)
            else:
                st.success(str(result))

    elif run_clicked and not question.strip():
        st.warning("Type a question first.")

    st.divider()

    # ── Step 3: PPT Insights ───────────────────────────────────────────────────
    st.markdown('<div class="section-label">Step 3 — PowerPoint Slide Suggestions</div>', unsafe_allow_html=True)
    st.caption("Phi-3 will analyze the data and draft 5 slide ideas for your marketing review deck.")

    if st.button("✨ Generate Slide Suggestions", type="primary", disabled=not ollama_ok):
        with st.spinner(f"Analyzing with {model}… (this may take ~30s)"):
            insights_text = get_insights(df, selected, model)
        st.session_state["insights"] = insights_text

    if "insights" in st.session_state:
        raw = st.session_state["insights"]

        # Split on "## Slide N" markers for cards
        slides = re.split(r"(?=##\s+Slide\s+\d+)", raw, flags=re.IGNORECASE)
        slides = [s.strip() for s in slides if s.strip()]

        if len(slides) >= 2:
            for slide in slides:
                st.markdown(f'<div class="insight-card">{slide}</div>', unsafe_allow_html=True)
        else:
            st.markdown(raw)

        st.download_button(
            label="⬇️ Download insights as .txt",
            data=raw,
            file_name=f"insights_{re.sub(r'[^\\w]', '_', selected)[:40]}.txt",
            mime="text/plain",
        )

    if not ollama_ok:
        st.info("Start Ollama to enable AI features: `ollama serve` then `ollama pull phi3`")


if __name__ == "__main__":
    main()
